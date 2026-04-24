import streamlit as st
import numpy_financial as npf
import pandas as pd
import altair as alt
from openpyxl import Workbook
from io import BytesIO


st.set_page_config(
    page_title="Business Case Tool",
    page_icon="💼",
    layout="wide"
)


def create_excel_export(
    project_name,
    investment,
    annual_savings,
    annual_extra_costs,
    annual_net_savings,
    years,
    discount_rate,
    roi,
    npv,
    payback_period,
    fcp_savings_year1,
    conclusion,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Case"

    ws["A1"] = "Business Case Tool"

    ws["A3"] = "Project"
    ws["B3"] = project_name

    ws["A5"] = "Initiële investering"
    ws["B5"] = investment

    ws["A6"] = "Jaarlijkse besparing"
    ws["B6"] = annual_savings

    ws["A7"] = "Jaarlijkse extra kosten"
    ws["B7"] = annual_extra_costs

    ws["A8"] = "Netto jaarlijkse besparing"
    ws["B8"] = annual_net_savings

    ws["A9"] = "Looptijd"
    ws["B9"] = years

    ws["A10"] = "Discontovoet"
    ws["B10"] = discount_rate

    ws["A12"] = "ROI"
    ws["B12"] = roi if roi is not None else "Niet berekenbaar"

    ws["A13"] = "NPV"
    ws["B13"] = npv

    ws["A14"] = "Terugverdientijd"
    ws["B14"] = payback_period if payback_period is not None else "Niet berekenbaar"

    ws["A15"] = "FCP Savings jaar 1"
    ws["B15"] = fcp_savings_year1

    ws["A17"] = "Conclusie"
    ws["B17"] = conclusion

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


st.title("Business Case Tool")

st.caption(
    "Onderbouwing voor concept- of sourcingwijzigingen binnen inkoop, inclusief NPV, ROI, terugverdientijd en FCP Savings jaar 1."
)

st.sidebar.header("Invoer")

project_name = st.sidebar.text_input("Naam business case")

investment = st.sidebar.number_input("Initiële investering (€)", value=0.0)
annual_savings = st.sidebar.number_input("Jaarlijkse besparing (€)", value=0.0)
annual_extra_costs = st.sidebar.number_input("Jaarlijkse extra kosten (€)", value=0.0)
years = st.sidebar.number_input("Looptijd (jaren)", value=1, step=1)
discount_rate_percent = st.sidebar.number_input("Discontovoet (%)", value=8.0)

discount_rate = discount_rate_percent / 100

st.write("---")

if investment <= 0:
    st.warning("Voer een investering groter dan 0 in")

if years <= 0:
    st.warning("Looptijd moet minimaal 1 jaar zijn")

if discount_rate_percent < 0:
    st.warning("Discontovoet kan niet negatief zijn")

if st.button("Bereken business case") and investment > 0 and years > 0:
    annual_net_savings = annual_savings - annual_extra_costs

    if investment != 0:
        total_benefit = annual_net_savings * years
        roi = (total_benefit - investment) / investment
    else:
        roi = None

    cashflows = [-investment] + [annual_net_savings] * int(years)
    npv = npf.npv(discount_rate, cashflows)

    npv_per_year = []
    cumulative_npv = 0

    for year, cashflow in enumerate(cashflows):
        discounted_cashflow = cashflow / ((1 + discount_rate) ** year)
        cumulative_npv += discounted_cashflow
        npv_per_year.append({
            "Jaar": year,
            "NPV": cumulative_npv
        })

    npv_df = pd.DataFrame(npv_per_year)

    if annual_net_savings > 0:
        payback_period = investment / annual_net_savings
    else:
        payback_period = None

    fcp_savings_year1 = annual_net_savings - investment

    st.header("Resultaten")

    col1, col2, col3, col4 = st.columns(4)

    col1.metric("Netto besparing / jaar", f"€ {annual_net_savings:,.0f}")

    if roi is not None:
        col2.metric("ROI", f"{roi:.1%}")
    else:
        col2.metric("ROI", "n.v.t.")

    col3.metric("NPV", f"€ {npv:,.0f}")

    if payback_period is not None:
        col4.metric("Payback", f"{payback_period:.1f} jr")
    else:
        col4.metric("Payback", "n.v.t.")

    st.write(f"FCP Savings jaar 1: € {fcp_savings_year1:,.0f}")

    st.write("---")
    
    st.write("---")
    
    st.header("NPV ontwikkeling per jaar")

    chart = alt.Chart(npv_df).mark_line(point=True).encode(
            x=alt.X("Jaar:O", title="Jaar", axis=alt.Axis(labelAngle=0)),
            y=alt.Y("NPV:Q", title="NPV (€)"),
            tooltip=["Jaar", "NPV"]
        ).properties(
            title="NPV ontwikkeling per jaar"
        )

    st.altair_chart(chart, use_container_width=True)

    st.header("Conclusie")

    if npv > 0 and (roi is not None and roi > 0):
        conclusion = "POSITIEVE BUSINESS CASE"
        st.success("POSITIEVE BUSINESS CASE ✅")
    elif npv > 0:
        conclusion = "TWEEDE CHECK AANBEVOLEN"
        st.warning("TWEEDE CHECK AANBEVOLEN ⚠️")
    else:
        conclusion = "NEGATIEVE BUSINESS CASE"
        st.error("NEGATIEVE BUSINESS CASE ❌")

    st.write("---")
    st.header("Export")

    excel_file = create_excel_export(
        project_name,
        investment,
        annual_savings,
        annual_extra_costs,
        annual_net_savings,
        years,
        discount_rate_percent,
        roi,
        npv,
        payback_period,
        fcp_savings_year1,
        conclusion,
    )

    st.download_button(
        label="Download business case als Excel",
        data=excel_file,
        file_name="business_case_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )